#coding=utf-8

import json
from openpyxl import load_workbook
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

def resolveJson(path):
    file = open(path, "rb")
    fileJson = json.load(file)
    futures = fileJson["obj"]["companies"]
    return (futures)

def wJsonExls(result,worksheet,fileIndex):
    rowIndex = 0
    for row in result:
        '''print "companyId=" + row["companyId"],\
              "companyName=" + row["companyName"],\
              "Lon="+row["companyLon"],\
              "Lat="+row["companyLat"]
        '''
        rowIndex = rowIndex + 1
        lenMainClm = len(row) # 主表列的数量
        lenMainAreaClm = len(row) + len(row["companyArea"]) # 主表列+区域列的数量
        # 总列宽
        # clumnLen = len(row) + len(row["companyArea"]) +  len(row["companyType"])
        # print clumnLen
        # print rowIndex
        clmIndex = 1
        for clm in row:
            # 每一行数据
            if 10 * fileIndex + rowIndex == 1 : # 记录表头
                if clm == "companyArea" :
                    caClmIndex = lenMainClm + 1 # companyArea 开始行
                    for ca in row[clm] :
                        d = worksheet.cell(row=10 * fileIndex + rowIndex, column=caClmIndex)
                        d.value = str(ca)
                        caClmIndex = caClmIndex + 1
                        # print ca,row[clm][ca]
                    # print "companyArea"
                elif clm == "companyType" : # 记录ctype
                    ctClmIndex = lenMainAreaClm + 1  # companyArea 开始行
                    for ct in row[clm] :
                        d = worksheet.cell(row=10 * fileIndex + rowIndex, column=ctClmIndex)
                        d.value = str(ct)
                        ctClmIndex = ctClmIndex + 1
                        # print ct,row[clm][ct]
                    # print "companyType"
                else:
                    d = worksheet.cell(row=10 * fileIndex + rowIndex, column=clmIndex)
                    d.value = str(clm)
            else : # 填充数据
                if clm == "companyArea" :
                    caClmIndex = lenMainClm + 1 # companyArea 开始行
                    for ca in row[clm] :
                        d = worksheet.cell(row=10 * fileIndex + rowIndex, column=caClmIndex)
                        d.value = str(row[clm][ca])
                        caClmIndex = caClmIndex + 1
                        # print ca,row[clm][ca]
                    # print "companyArea"
                elif clm == "companyType" : # 记录ctype
                    ctClmIndex = lenMainAreaClm + 1  # companyArea 开始行
                    for ct in row[clm] :
                        d = worksheet.cell(row=10 * fileIndex + rowIndex, column=ctClmIndex)
                        d.value = str(row[clm][ct])
                        ctClmIndex = ctClmIndex + 1
                        # print ct,row[clm][ct]
                    # print "companyType"
                else:
                    d = worksheet.cell(row=10 * fileIndex + rowIndex, column=clmIndex)
                    d.value = str(row[clm])
                # print clmIndex
                #d = worksheet.cell(row=rowIndex, column=clmIndex)
                #print clm, row[clm]
                #d.value = str(row[clm])
            # print clmIndex
            clmIndex = clmIndex + 1
    return worksheet

def output():
    result = resolveJson(path)
    # print(result)
    # 创建一个空的xlxs工作簿文件
    workbook = Workbook()
    workbook.save(filename="e:\\cmp.xlsx")
    worksheet = workbook.get_sheet_by_name("Sheet")
    rowIndex = 0
    for row in result:
        '''print "companyId=" + row["companyId"],\
              "companyName=" + row["companyName"],\
              "Lon="+row["companyLon"],\
              "Lat="+row["companyLat"]
        '''
        rowIndex = rowIndex + 1
        lenMainClm = len(row) # 主表列的数量
        lenMainAreaClm = len(row) + len(row["companyArea"]) # 主表列+区域列的数量
        # 总列宽
        # clumnLen = len(row) + len(row["companyArea"]) +  len(row["companyType"])
        # print clumnLen
        # print rowIndex
        clmIndex = 1
        for clm in row:
            # 每一行数据
            if rowIndex == 1 : # 记录表头
                if clm == "companyArea" :
                    caClmIndex = lenMainClm + 1 # companyArea 开始行
                    for ca in row[clm] :
                        d = worksheet.cell(row=rowIndex, column=caClmIndex)
                        d.value = str(ca)
                        caClmIndex = caClmIndex + 1
                        # print ca,row[clm][ca]
                    # print "companyArea"
                elif clm == "companyType" : # 记录ctype
                    ctClmIndex = lenMainAreaClm + 1  # companyArea 开始行
                    for ct in row[clm] :
                        d = worksheet.cell(row=rowIndex, column=ctClmIndex)
                        d.value = str(ct)
                        ctClmIndex = ctClmIndex + 1
                        # print ct,row[clm][ct]
                    # print "companyType"
                else:
                    d = worksheet.cell(row=rowIndex, column=clmIndex)
                    d.value = str(clm)
            else : # 填充数据
                if clm == "companyArea" :
                    caClmIndex = lenMainClm + 1 # companyArea 开始行
                    for ca in row[clm] :
                        d = worksheet.cell(row=rowIndex, column=caClmIndex)
                        d.value = str(row[clm][ca])
                        caClmIndex = caClmIndex + 1
                        # print ca,row[clm][ca]
                    # print "companyArea"
                elif clm == "companyType" : # 记录ctype
                    ctClmIndex = lenMainAreaClm + 1  # companyArea 开始行
                    for ct in row[clm] :
                        d = worksheet.cell(row=rowIndex, column=ctClmIndex)
                        d.value = str(row[clm][ct])
                        ctClmIndex = ctClmIndex + 1
                        # print ct,row[clm][ct]
                    # print "companyType"
                else:
                    d = worksheet.cell(row=rowIndex, column=clmIndex)
                    d.value = str(row[clm])
                # print clmIndex
                #d = worksheet.cell(row=rowIndex, column=clmIndex)
                #print clm, row[clm]
                #d.value = str(row[clm])
            # print clmIndex
            clmIndex = clmIndex + 1

    print "------------------------"
    workbook.save(filename="e:\\cmp.xlsx")

if __name__ == '__main__':
    workbook = Workbook()
    workbook.save(filename="e:\\cmp.xlsx")
    worksheet = workbook.get_sheet_by_name("Sheet")
    fileIndex = 0

    for i in range(0, 102):
        result = resolveJson("E:\\tjsfjg\\" + str(i) + ".json")
        worksheet = wJsonExls(result,worksheet,i)

    workbook.save(filename="e:\\cmp.xlsx")
# path = r"E:\tjsfjg\1.json"
# output()
